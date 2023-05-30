import re
import time
import sqlite3
import psutil
import openpyxl
import win32evtlog
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import QApplication, QAction, QMessageBox
from PyQt5.QtGui import QIcon
from scapy.all import sniff
from scapy.layers.inet import TCP
from threading import Thread
from main import LoginPage


# Function to create the database to store the audit data (log, resources and analysed data)
def create_audit_database():
    conn = sqlite3.connect("audit_data.db")
    cursor = conn.cursor()

    cursor.execute('''CREATE TABLE IF NOT EXISTS log_data (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      timestamp TEXT,
                      log_entry TEXT);''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS resource_data (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      timestamp TEXT,
                      resource_data TEXT);''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS analysis_results (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      timestamp TEXT,
                      result TEXT,
                      analyser TEXT);''')

    # Committing and close the cursor
    cursor.close()
    
    # Returning the connection instead of closing it
    return conn


# Function to system log monitoring
def monitor_logs(conn):
    # Create a handle to the system log.
    handler = win32evtlog.OpenEventLog(None, "System")

    # Reading the most recent log record
    flags = win32evtlog.EVENTLOG_BACKWARDS_READ | win32evtlog.EVENTLOG_SEQUENTIAL_READ
    events = win32evtlog.ReadEventLog(handler, flags, 0)
    log_entry = events[0].StringInserts if events else "No new log entries"
    log_entry = log_entry[0]
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    cursor = conn.cursor()
    cursor.execute("INSERT INTO log_data (timestamp, log_entry) VALUES (?, ?)", (timestamp, log_entry))
    conn.commit()

    return log_entry


# Function to system resource monitoring
def monitor_system_resources(conn):
    resource_data = {
        'cpu_percent': psutil.cpu_percent(),
        'memory_percent': psutil.virtual_memory().percent,
        'disk_percent': psutil.disk_usage('/').percent
    }
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO resource_data (timestamp, resource_data) VALUES (?, ?)",
                   (timestamp, str(resource_data)))
    conn.commit()
    return resource_data


# Function to system log analysing
def analyse_logs(log_entry, conn):
    # Defining patterns to look for
    patterns = [r'Error', r'Critical', r'Fatal', r'Failed Login Attempt']

    # Flag for whether an intrusion is detected
    intrusion_detected = False

    for pattern in patterns:
        if re.search(pattern, log_entry, re.IGNORECASE):
            intrusion_detected = True
            break

    result = "Intrusion detected" if intrusion_detected else "No intrusions detected"
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    cursor = conn.cursor()
    cursor.execute("INSERT INTO analysis_results (timestamp, result, analyser) VALUES (?, ?, ?)",
                   (timestamp, result, "log_analyser"))
    conn.commit()

    return intrusion_detected


# Function to system resource analysing
def analyse_system_resources(resource_data, conn):
    # if CPU usage or memory usage is over 90 (%), an intrusion is detected.
    cpu_threshold = 90
    memory_threshold = 90
    intrusion_detected = resource_data['cpu_percent'] > cpu_threshold or resource_data['memory_percent'] > memory_threshold
    result = "Intrusion detected" if intrusion_detected else "No intrusion detected"
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    cursor = conn.cursor()
    cursor.execute("INSERT INTO analysis_results (timestamp, result, analyser) VALUES (?, ?, ?)",
                   (timestamp, result, "system_resources_analyser"))
    conn.commit()

    return intrusion_detected


# Function to create Active Response Unit (retrieving the analysed data and check unusual behaviors)
def active_response_unit(conn, alert_callback, should_stop):
    while not should_stop():
        # Retrieving data of the monitoring (system logs and resources)
        log_entry = monitor_logs(conn)
        resource_data = monitor_system_resources(conn)

        # Analysing data which receieved
        log_intrusion_detected = analyse_logs(log_entry, conn)
        resource_intrusion_detected = analyse_system_resources(resource_data, conn)

        # Monitoring network traffic in a separate thread (using a thread to do the procedure continuously without blocking the other operations)
        network_traffic_thread = Thread(target=monitor_network_traffic, args=(analyse_packet, alert_callback))
        network_traffic_thread.start()

        # Displaying the alert when instrusion is detected
        if log_intrusion_detected or resource_intrusion_detected:
            alert_callback()

        # Waiting time between checkings (in seconds)
        time.sleep(2)


# Function to monitor network traffic
def monitor_network_traffic(packet_callback, alert_callback):
    def handle_packet(packet):
        if packet_callback(packet):
            alert_callback()

    sniff(filter="ip", prn=handle_packet) # Capturing IP packets from the network


# Function to detect abnormal network traffic
def analyse_packet(packet):
    # if the packet is TCP and the data size is over 5000 bytes, considering it as abnormal.
    tcp_data_threshold = 5000
    abnormal_traffic_detected = TCP in packet and len(packet[TCP].payload) > tcp_data_threshold

    return abnormal_traffic_detected


# Function to export all the data to an Excel file
def export_data_to_excel(conn):
    # Creating a workbook and add sheets for system log and resource monitoring, and analysis_results
    workbook = openpyxl.Workbook()

    log_data_sheet = workbook.active
    log_data_sheet.title = "log_data"
    resource_data_sheet = workbook.create_sheet("resource_data")
    analysis_results_sheet = workbook.create_sheet("analysis_results")

    # Writing log data to the log_data_sheet (sheet 01)
    cursor = conn.cursor()
    cursor.execute("SELECT id, timestamp, log_entry FROM log_data ORDER BY id ASC")
    log_data = cursor.fetchall()

    log_data_sheet.append(["ID", "Timestamp", "Log Entry"])
    for row in log_data:
        log_data_sheet.append(row)

    # Writing system resources data to the resource_data_sheet (sheet 02)
    cursor.execute("SELECT id, timestamp, resource_data FROM resource_data ORDER BY id ASC")
    resource_data = cursor.fetchall()

    resource_data_sheet.append(["ID", "Timestamp", "Resource Data"])
    for row in resource_data:
        resource_data_sheet.append(row)

    # Writing analysis results to the analysis_results_sheet (sheet 03)
    cursor.execute("SELECT id, timestamp, result, analyser FROM analysis_results ORDER BY id ASC")
    analysis_results = cursor.fetchall()

    analysis_results_sheet.append(["ID", "Timestamp", "Result", "Analyser"])
    for row in analysis_results:
        analysis_results_sheet.append(row)

    # Protecting the workbook for view only (to prevent from changing the data)
    log_data_sheet.protection.sheet = True
    resource_data_sheet.protection.sheet = True
    analysis_results_sheet.protection.sheet = True
    workbook.security.lockStructure = True

    # Saving the workbook to the Excel file
    workbook.save("data_export.xlsx")


# Function to confirm exporting data to the Excel file (confirmation box)
def show_confirmation_dialog(conn):
    confirmation_box = QMessageBox()
    confirmation_box.setIcon(QMessageBox.Question)
    confirmation_box.setWindowTitle("Export all the data")
    confirmation_box.setText("Are you sure you want to export all the data?")
    confirmation_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    confirmation_box.setDefaultButton(QMessageBox.No)

    # Adding the TrueIDS logo to the confirmation box's title bar
    logo_icon = QIcon("logo.png")
    confirmation_box.setWindowIcon(logo_icon)

    # Exporting data to the Excel file if the user clicks "Yes"
    result = confirmation_box.exec_()
    if result == QMessageBox.Yes:
        export_data_to_excel(conn)



# Class for the home page of the TrueIDS applciation
class homePage(QtWidgets.QMainWindow):
    def __init__(self, conn, parent=None):
        super().__init__(parent)
        self.conn = conn
        
        # Adding the title of the windows, and the window position with size
        self.setWindowTitle("TrueIDS")
        self.setGeometry(100, 100, 800, 600)

        # Making the window position to the center on the screen
        screen_geometry = QtWidgets.QDesktopWidget().availableGeometry()
        window_geometry = self.geometry()
        x = (screen_geometry.width() - window_geometry.width()) // 2
        y = (screen_geometry.height() - window_geometry.height()) // 2
        self.move(x, y)

        # Adding the TrueIDS logo to the titlebar
        icon = QtGui.QIcon("logo.png")
        self.setWindowIcon(icon)


        # Creating the menu bar
        menu_bar = self.menuBar()

        # "File" menu
        file_menu = menu_bar.addMenu("File")
        display_all_data_action = file_menu.addAction("Display all the data") # First menu item in "File" menu
        display_all_data_action.triggered.connect(self.display_all_data)

        export_all_data_action = QAction("Export all the data", self) # Second menu item in "File" menu
        export_all_data_action.triggered.connect(lambda: show_confirmation_dialog(self.conn))
        file_menu.addAction(export_all_data_action)

        logout_action = QAction('Logout', self) # Third menu item in "File" menu
        logout_action.triggered.connect(self.show_logout_confirmation_dialog)
        file_menu.addAction(logout_action)

        exit_action = QAction('Exit', self) # Fourth menu item in "File" menu
        exit_action.triggered.connect(self.show_exit_confirmation_dialog)
        file_menu.addAction(exit_action)

        # "Edit" menu [will be adding two items to create new users and change PWs of users (this whole menu will only be visible to the system admin.)]
        edit_menu = menu_bar.addMenu("Edit")
        
        # "Preferences" menu
        pref_menu = menu_bar.addMenu("Preferences")

        # "Help" menu
        help_menu = menu_bar.addMenu("Help")
        view_details_action = help_menu.addAction("About TrueIDS") # First menu item in "Help" menu
        view_details_action.triggered.connect(self.view_application_details)


        # Creating the status bar
        status_bar = self.statusBar()
        status_bar.showMessage("Ready")

        # Creating the vertical layout to add widgets for the home page
        self.main_layout = QtWidgets.QVBoxLayout()

        # Creating the button to start monitoring
        self.start_button = QtWidgets.QPushButton("Start Monitoring")
        self.start_button.clicked.connect(self.start_monitoring)
        self.main_layout.addWidget(self.start_button)

        # Creating the button to stop monitoring
        self.stop_button = QtWidgets.QPushButton("Stop Monitoring")
        self.stop_button.clicked.connect(self.stop_monitoring)
        self.stop_button.setEnabled(False)
        self.main_layout.addWidget(self.stop_button)

        # Creating the button to display the results from the last stopped session
        self.last_session_button = QtWidgets.QPushButton("View Last Session Results")
        self.last_session_button.clicked.connect(self.view_last_session_results)
        self.main_layout.addWidget(self.last_session_button)

        # Creating the text area to display intrusion alerts and data/ results
        self.alert_text_area = QtWidgets.QTextEdit(self)
        self.main_layout.addWidget(self.alert_text_area)

        # Creating the central widget to hold the layout and widgets
        central_widget = QtWidgets.QWidget(self)
        central_widget.setLayout(self.main_layout)
        self.setCentralWidget(central_widget)

        # Starting the Active Response Unit in a separate thread (to avoid blocking the processing of the main thread of the application)
        self.active_response_thread = QtCore.QThread(self)
        self.active_response_worker = ActiveResponseWorker(self.display_intrusion_alert)
        self.active_response_worker.moveToThread(self.active_response_thread)
        self.active_response_thread.started.connect(self.active_response_worker.run)


    # Function to display the "About TrueIDS" menu item (information about the application)
    def view_application_details(self):
        about_dialog = AboutDialog(self)
        about_dialog.exec_()


    # Function to confirm logging out from the home page
    def show_logout_confirmation_dialog(self):
        confirmation_box = QMessageBox()
        confirmation_box.setIcon(QMessageBox.Question)
        confirmation_box.setWindowTitle("Logout")
        confirmation_box.setText("Are you sure you want to logout?")
        confirmation_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        confirmation_box.setDefaultButton(QMessageBox.No)

        # Adding the TrueIDS logo to the confirmation box's title bar
        logo_icon = QIcon("logo.png")
        confirmation_box.setWindowIcon(logo_icon)

        # Logging out from the application if the user clicks "Yes"
        result = confirmation_box.exec_()
        if result == QMessageBox.Yes:
            self.logout()

    
    # Funtion to redirect to the login window
    def logout(self):
        self.hide() # hide the current window
        self.loginPage = LoginPage() # create a new instance of LoginPage
        self.loginPage.show() # show the LoginPage


    # Function to confirm exiting the application
    def show_exit_confirmation_dialog(self):
        confirmation_box = QMessageBox()
        confirmation_box.setIcon(QMessageBox.Question)
        confirmation_box.setWindowTitle("Exit")
        confirmation_box.setText("Are you sure you want to exit?")
        confirmation_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        confirmation_box.setDefaultButton(QMessageBox.No)

        # Adding the TrueIDS logo to the confirmation box's title bar
        logo_icon = QIcon("logo.png")
        confirmation_box.setWindowIcon(logo_icon)

        # Exiting from the application if the user clicks "Yes"
        result = confirmation_box.exec_()
        if result == QMessageBox.Yes:
            QApplication.quit()


    # Function to start the monitoring process
    def start_monitoring(self):
        self.alert_text_area.append("\n\nMonitoring started at {}.\n".format(time.strftime("%Y-%m-%d %H:%M:%S"))) # Showing the monitoring has started in the textarea with the date and time
        self.start_button.setEnabled(False) # Setting the "Start Monnitoring" button disabled
        self.stop_button.setEnabled(True) # Setting the "Stop Monnitoring" button enabled
        self.active_response_worker._should_stop = False # Setting the worker to keep continue the monitoring
        self.active_response_thread.start() # Running the active response worker in a separate thread


    # Function to stop the monitoring process
    def stop_monitoring(self):
        self.alert_text_area.append("\nMonitoring stopped at {}.\n".format(time.strftime("%Y-%m-%d %H:%M:%S"))) # Show the monitoring has stopped in the textarea with the date and time
        self.active_response_worker.stop() # Setting the worker to stop the monitoring
        self.active_response_thread.quit() # Setting the separate thread to exit
        self.active_response_thread.wait() # Setting the separate thread to block the monitoring process until the thread terminates
        self.start_button.setEnabled(True) # Setting the "Start Monnitoring" button enabled again
        self.stop_button.setEnabled(False) # Setting the "Stop Monnitoring" button disabled again


    # Function to view the data/results of the last monitoring session
    def view_last_session_results(self):
        # Implementing the logic to display the data from the last session
        cursor = self.conn.cursor()

        # Fetching log_data
        cursor.execute("SELECT timestamp, log_entry FROM log_data ORDER BY id DESC LIMIT 1")
        log_data = cursor.fetchone()
        if log_data:
            timestamp, log_entry = log_data
            self.alert_text_area.append("\n\n\nLast log entry at {}:\n{}".format(timestamp, log_entry))

        # Fetching resource_data
        cursor.execute("SELECT timestamp, resource_data FROM resource_data ORDER BY id DESC LIMIT 1")
        resource_data = cursor.fetchone()
        if resource_data:
            timestamp, resource_data_str = resource_data
            resource_data = eval(resource_data_str)
            self.alert_text_area.append("\nLast resource data at {}:\n{}".format(timestamp, resource_data))

        # Fetching analysis_results
        cursor.execute("SELECT timestamp, result, analyser FROM analysis_results ORDER BY id DESC LIMIT 2")
        analysis_results = cursor.fetchall()
        for result in analysis_results:
            timestamp, result, analyser = result
            self.alert_text_area.append("\nLast analysis result from {} at {}:\n{}".format(analyser, timestamp, result))


    # Function to display all the data in the textarea of the home page
    def display_all_data(self):
        # Implementing the logic to display all the data
        cursor = self.conn.cursor()

        # Fetching log_data
        cursor.execute("SELECT id, timestamp, log_entry FROM log_data ORDER BY id")
        #cursor.execute("SELECT id, timestamp, log_entry FROM log_data ORDER BY id DESC LIMIT 1")
        log_data = cursor.fetchall()
        self.alert_text_area.append("\n\n\nAll log entries:")
        for entry in log_data:
            id, timestamp, log_entry = entry
            self.alert_text_area.append("{} - {}: {}".format(id, timestamp, log_entry))

        # Fetching resource_data
        cursor.execute("SELECT id, timestamp, resource_data FROM resource_data ORDER BY id")
        resource_data = cursor.fetchall()
        self.alert_text_area.append("\n\nAll resource data:")
        for data in resource_data:
            id, timestamp, resource_data_str = data
            resource_data = eval(resource_data_str)
            self.alert_text_area.append("{} - {}: {}".format(id, timestamp, resource_data))

        # Fetching analysis_results
        cursor.execute("SELECT id, timestamp, result, analyser FROM analysis_results ORDER BY id")
        analysis_results = cursor.fetchall()
        self.alert_text_area.append("\n\nAll analysis results:")
        for result in analysis_results:
            id, timestamp, result, analyser = result
            self.alert_text_area.append("{} - From {} at {}: {}".format(id, analyser, timestamp, result))


    # Function to display instrusion alerts and data/results
    def display_intrusion_alert(self):
        # Check if monitoring should stop before creating an alert message
        if not self.active_response_worker._should_stop:
            # Creating an alert message
            alert_message = "Intrusion detected at {}!".format(time.strftime("%Y-%m-%d %H:%M:%S"))

            # Displaying the alert in the text area
            self.alert_text_area.append(alert_message)



# Class to show the information about the TrueIDS in a dialog box (QDialog)
class AboutDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)

        # Title of the window
        self.setWindowTitle("About TrueIDS")
        self.setFixedSize(500, 400)

        # A vertical layout to arange the widgets in a column
        layout = QtWidgets.QVBoxLayout()

        # Adding the TrueIDS logo in the dialog box
        logo_label = QtWidgets.QLabel()
        logo_pixmap = QtGui.QPixmap("logo.png")

        # Adusting the size of the logo (width x height)
        scaled_logo_pixmap = logo_pixmap.scaled(100, 100, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        logo_label.setPixmap(scaled_logo_pixmap)
        logo_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(logo_label)

        # Adding a text ("TreuIDS") in a label
        app_title_label = QtWidgets.QLabel("TrueIDS")
        app_title_label.setAlignment(QtCore.Qt.AlignCenter)
        app_title_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        layout.addWidget(app_title_label)

        # Adding the application version
        app_version_label = QtWidgets.QLabel("Version: 1.0.0")
        app_version_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(app_version_label)

        # Adding the application description (What TrueIDS is)
        app_description_label = QtWidgets.QLabel("TrueIDS is an Intrusion Detection System that allows to analyse system logs, resources and moniter the network traffic to detect intruders while auditing necessary data to a database.")
        app_description_label.setWordWrap(True)
        app_description_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(app_description_label)

        # Adding the application description (copyright information)
        app_description_label = QtWidgets.QLabel("Copyright © 2022-2023 by Vehan Rathnayake (Computer Security Undergraduate at University of Plymouth, UK)")
        app_description_label.setWordWrap(True)
        app_description_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(app_description_label)

        # Adding an "OK" button to close the dialog box
        ok_button = QtWidgets.QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        layout.addWidget(ok_button)

        # Setting the layout for the "About TrueIDS" dialog box
        self.setLayout(layout)



# Class to handle the execution of the active response unit function in a separate thread
class ActiveResponseWorker(QtCore.QObject):
    def __init__(self, alert_callback):
        super().__init__()
        self._should_stop = False
        self.alert_callback = alert_callback
        

    @QtCore.pyqtSlot()
    # Creating a new SQLite connection and execute the the active response unit function with the given callback and the stopping condition
    def run(self):
        conn = create_audit_database()  # Create a new SQLite connection here
        active_response_unit(conn, self.alert_callback, self.should_stop)


    # Returning the current stopping status of the worker
    def should_stop(self):
        return self._should_stop


    # Setting the stopping status of the worker to "True" to stop the execution of the active response unit function
    def stop(self):
        self._should_stop = True



#if __name__ == "__main__":
    #conn = create_audit_database()
    #app = QtWidgets.QApplication(sys.argv)
    #window = homePage(conn)
    #window.show()
    #sys.exit(app.exec_())